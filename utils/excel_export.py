import os
import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

def export_attendance_to_excel(db_connection, filename="attendance_report.xlsx"):
    """
    Exports attendance data from SQLite database to a styled Excel sheet using Pandas & openpyxl.
    """
    # Fetch all records
    df = pd.read_sql_query('''
        SELECT student_name, branch, section, crn, urn, phone, is_nss_volunteer, event, attendance_mode, created_at
        FROM attendance
        ORDER BY id DESC
    ''', db_connection)
    
    # Rename columns for a clean presentation
    df.columns = [
        'Student Name', 
        'Branch', 
        'Section',
        'Class Roll No (CRN)', 
        'University Roll No (URN)', 
        'Phone Number', 
        'NSS Volunteer',
        'Event Name',
        'Attendance Mode',
        'Date & Time'
    ]

    os.makedirs('excel_reports', exist_ok=True)
    filepath = os.path.join('excel_reports', filename)

    # Save Excel
    with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='NSS Attendance')
        
        # Access openpyxl sheet to apply high-quality styling
        workbook = writer.book
        worksheet = writer.sheets['NSS Attendance']
        
        # Theme colors: Navy Blue primary for headers, light gray borders
        header_font = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='002147', end_color='002147', fill_type='solid') # Navy Blue
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_align = Alignment(horizontal='left', vertical='center')
        
        thin_border = Border(
            left=Side(style='thin', color='E2E8F0'),
            right=Side(style='thin', color='E2E8F0'),
            top=Side(style='thin', color='E2E8F0'),
            bottom=Side(style='thin', color='E2E8F0')
        )
        
        # Style headers
        for col_idx in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
            
        worksheet.row_dimensions[1].height = 28
        
        # Style data rows
        data_font = Font(name='Segoe UI', size=10)
        volunteer_fill_yes = PatternFill(start_color='E6F4EA', end_color='E6F4EA', fill_type='solid') # soft green
        volunteer_fill_no = PatternFill(start_color='FCE8E6', end_color='FCE8E6', fill_type='solid') # soft red
        
        for row_idx in range(2, len(df) + 2):
            worksheet.row_dimensions[row_idx].height = 20
            for col_idx in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.font = data_font
                cell.border = thin_border
                
                # Apply conditional alignments
                val = str(cell.value or '')
                # Align numbers and dates to center, names/events to left
                if col_idx in [3, 4, 5, 6, 7, 9, 10]:
                    cell.alignment = center_align
                else:
                    cell.alignment = left_align
                    
                # Style volunteer column (col_idx = 7 is NSS Volunteer)
                if col_idx == 7:
                    if val.strip().lower() in ['yes', 'y']:
                        cell.fill = volunteer_fill_yes
                    elif val.strip().lower() in ['no', 'n']:
                        cell.fill = volunteer_fill_no

        # Auto-fit columns with padding
        for col in worksheet.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                val = str(cell.value or '')
                if len(val) > max_len:
                    max_len = len(val)
            worksheet.column_dimensions[col_letter].width = max(max_len + 4, 12)

    return filepath
