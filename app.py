import os
import random
import pandas as pd
from functools import wraps

from flask import (
    Flask, render_template, request, redirect,
    url_for, flash, jsonify, send_file, session
)
import sqlite3

from utils.qr_generator import generate_event_qr
from utils.excel_export import export_attendance_to_excel
from utils.proxy_detector import detect_proxies
from utils.ocr_reader import OCRReader

# ─────────────────────────────────────────────
app = Flask(__name__)
app.secret_key = "nss_secret_key"

DB_PATH   = "database.db"
OCR_EXCEL = "uploads/attendance.xlsx"
ocr_reader = OCRReader(OCR_EXCEL)


# ─────────────────────────────────────────────
#  Helpers
# ─────────────────────────────────────────────

def get_db_connection():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def admin_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        if not session.get("is_admin"):
            return redirect(url_for("admin_login"))
        return f(*args, **kwargs)
    return wrapper


def migrate_db():
    """Safely add location columns if they don't exist yet."""
    conn = get_db_connection()
    existing = [row[1] for row in conn.execute("PRAGMA table_info(attendance)").fetchall()]
    for col, coltype in [("latitude", "TEXT"), ("longitude", "TEXT"), ("location_status", "TEXT")]:
        if col not in existing:
            conn.execute(f"ALTER TABLE attendance ADD COLUMN {col} {coltype}")
    conn.commit()
    conn.close()


# ─────────────────────────────────────────────
#  Public routes
# ─────────────────────────────────────────────

@app.route("/")
def index():
    return render_template("index.html")


# ── OTP ──────────────────────────────────────

@app.route("/send_otp", methods=["POST"])
def send_otp():
    phone = request.form.get("phone", "").strip()

    if not phone:
        return jsonify({"status": "error", "message": "Phone number required"}), 400

    if not phone.isdigit() or len(phone) != 10:
        return jsonify({"status": "error", "message": "Enter a valid 10-digit phone number"}), 400

    otp = random.randint(1000, 9999)
    session["otp"]      = str(otp)
    session["phone"]    = phone
    session["verified"] = False

    print("OTP SENT:", otp)   # Replace with real SMS gateway in production

    return jsonify({
        "status":  "success",
        "message": "OTP generated successfully",
        "otp":     str(otp),   # Remove this line in production
    })


@app.route("/verify_otp", methods=["POST"])
def verify_otp():
    user_otp = request.form.get("otp", "").strip()

    if not session.get("otp"):
        return jsonify({"message": "Please request OTP first"}), 400

    if user_otp == session.get("otp"):
        session["verified"] = True
        return jsonify({"message": "OTP verified"})

    session["verified"] = False
    return jsonify({"message": "Invalid OTP"}), 400


# ── Student form ──────────────────────────────

@app.route("/student_form", methods=["GET", "POST"])
def student_form():
    conn = get_db_connection()

    if request.method == "POST":

        if not session.get("verified"):
            flash("Please verify OTP first", "error")
            conn.close()
            return redirect(request.url)

        student_name     = request.form.get("student_name",     "").strip()
        branch           = request.form.get("branch",           "").strip()
        section          = request.form.get("section",          "").strip()
        crn              = request.form.get("crn",              "").strip()
        urn              = request.form.get("urn",              "").strip()
        phone            = request.form.get("phone",            "").strip()
        is_nss_volunteer = request.form.get("is_nss_volunteer", "").strip()
        event            = request.form.get("event",            "").strip()
        latitude         = request.form.get("latitude",         "").strip()
        longitude        = request.form.get("longitude",        "").strip()
        location_status  = request.form.get("location_status",  "unknown").strip()

        if not all([student_name, branch, section, crn, urn, phone, is_nss_volunteer, event]):
            flash("All fields are required", "error")
            conn.close()
            return redirect(url_for("student_form", event=event))

        # Block submission if location not captured
        if location_status != "captured" or not latitude or not longitude:
            flash("Location is required. Please allow location access and try again.", "error")
            conn.close()
            return redirect(url_for("student_form", event=event))

        try:
            existing = conn.execute(
                "SELECT * FROM attendance WHERE (urn = ? OR phone = ?) AND event = ?",
                (urn, phone, event)
            ).fetchone()

            if existing:
                flash("Attendance already marked", "error")
                return redirect(url_for("student_form", event=event))

            conn.execute("""
                INSERT INTO attendance
                  (student_name, branch, section, crn, urn, phone,
                   is_nss_volunteer, event, attendance_mode,
                   latitude, longitude, location_status)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                student_name, branch, section, crn, urn, phone,
                is_nss_volunteer, event, "QR Scan",
                latitude, longitude, location_status
            ))

            conn.commit()
            session.pop("otp",      None)
            session.pop("verified", None)
            flash("Attendance marked successfully! ✓", "success")

        finally:
            conn.close()

        return redirect(url_for("student_form", event=event))

    # GET
    event_name = request.args.get("event", "").strip()
    events = conn.execute("""
        SELECT DISTINCT event FROM attendance
        WHERE event IS NOT NULL AND event != ''
        ORDER BY event ASC
    """).fetchall()
    conn.close()

    return render_template(
        "student_form.html",
        event_name=event_name,
        events_list=[row["event"] for row in events],
    )


# ─────────────────────────────────────────────
#  Admin
# ─────────────────────────────────────────────

ADMIN_USERNAME = "nss_admin_2026"
ADMIN_PASSWORD = "Arsh123"


@app.route("/admin/login", methods=["GET", "POST"])
def admin_login():
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "").strip()

        if username == ADMIN_USERNAME and password == ADMIN_PASSWORD:
            session["is_admin"] = True
            return redirect(url_for("admin"))

        flash("Invalid credentials", "error")

    return render_template("admin_login.html")


@app.route("/admin/logout")
def admin_logout():
    session.clear()
    return redirect(url_for("admin_login"))


@app.route("/admin")
@admin_required
def admin():
    conn = get_db_connection()

    total_records      = conn.execute("SELECT COUNT(*) FROM attendance").fetchone()[0]
    suspicious_records = detect_proxies(conn)

    branch_stats = conn.execute("""
        SELECT branch, COUNT(*) as count FROM attendance
        GROUP BY branch ORDER BY count DESC
    """).fetchall()

    event_stats = conn.execute("""
        SELECT event, COUNT(*) as count FROM attendance
        GROUP BY event ORDER BY count DESC
    """).fetchall()

    # Students who submitted location
    location_records = conn.execute("""
        SELECT id, student_name, crn, urn, branch, event,
               latitude, longitude, location_status
        FROM attendance
        WHERE latitude IS NOT NULL AND latitude != ''
        ORDER BY id DESC
    """).fetchall()

    conn.close()

    stats = {
        "total_records":     total_records,
        "active_events":     len(os.listdir("static/qr_codes")) if os.path.exists("static/qr_codes") else 0,
        "suspicious_alerts": len(suspicious_records),
        "location_count":    len(location_records),
    }

    return render_template(
        "admin_dashboard.html",
        stats=stats,
        suspicious_records=suspicious_records,
        branch_stats=branch_stats,
        event_stats=event_stats,
        location_records=location_records,
        qr_generated=False,
    )


# ── Manual entry ──────────────────────────────

@app.route("/admin/manual_entry", methods=["POST"])
@admin_required
def manual_entry():
    student_name     = request.form.get("student_name",     "").strip()
    branch           = request.form.get("branch",           "").strip()
    section          = request.form.get("section",          "").strip()
    crn              = request.form.get("crn",              "").strip()
    urn              = request.form.get("urn",              "").strip()
    phone            = request.form.get("phone",            "").strip()
    is_nss_volunteer = request.form.get("is_nss_volunteer", "").strip()
    event            = request.form.get("event",            "").strip()

    conn = get_db_connection()
    try:
        existing = conn.execute(
            "SELECT * FROM attendance WHERE (urn = ? OR phone = ?) AND event = ?",
            (urn, phone, event)
        ).fetchone()

        if existing:
            flash("Already marked", "error")
            return redirect(url_for("admin"))

        conn.execute("""
            INSERT INTO attendance
              (student_name, branch, section, crn, urn, phone,
               is_nss_volunteer, event, attendance_mode,
               latitude, longitude, location_status)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (student_name, branch, section, crn, urn, phone,
              is_nss_volunteer, event, "Manual",
              None, None, "manual_entry"))

        conn.commit()
        flash("Saved successfully", "success")

    finally:
        conn.close()

    return redirect(url_for("admin"))


# ── QR code generation ────────────────────────

@app.route("/admin/generate_qr", methods=["POST"])
@admin_required
def generate_qr():
    event_name = request.form.get("event_name", "").strip()

    if not event_name:
        flash("Event name required", "error")
        return redirect(url_for("admin"))

    try:
        filename, qr_link = generate_event_qr(event_name, request.host_url)
        conn = get_db_connection()

        total_records      = conn.execute("SELECT COUNT(*) FROM attendance").fetchone()[0]
        suspicious_records = detect_proxies(conn)
        branch_stats       = conn.execute("SELECT branch, COUNT(*) as count FROM attendance GROUP BY branch ORDER BY count DESC").fetchall()
        event_stats        = conn.execute("SELECT event, COUNT(*) as count FROM attendance GROUP BY event ORDER BY count DESC").fetchall()
        location_records   = conn.execute("""
            SELECT id, student_name, crn, urn, branch, event,
                   latitude, longitude, location_status
            FROM attendance
            WHERE latitude IS NOT NULL AND latitude != ''
            ORDER BY id DESC
        """).fetchall()
        conn.close()

        stats = {
            "total_records":     total_records,
            "active_events":     len(os.listdir("static/qr_codes")) if os.path.exists("static/qr_codes") else 0,
            "suspicious_alerts": len(suspicious_records),
            "location_count":    len(location_records),
        }

        flash("QR generated successfully!", "success")
        return render_template(
            "admin_dashboard.html",
            stats=stats,
            suspicious_records=suspicious_records,
            branch_stats=branch_stats,
            event_stats=event_stats,
            location_records=location_records,
            qr_generated=True,
            qr_filename=filename,
            qr_link=qr_link,
        )

    except Exception as e:
        print("QR ERROR:", e)
        flash(f"QR generation failed: {str(e)}", "error")
        return redirect(url_for("admin"))


# ── OCR routes ────────────────────────────────

@app.route("/admin/ocr_upload", methods=["POST"])
@admin_required
def ocr_upload():
    """Preview: upload manual sheet image → show comparison table."""
    image = request.files.get("manual_sheet")
    if not image or image.filename == "":
        flash("Please select an image", "error")
        return redirect(url_for("admin"))

    os.makedirs("uploads", exist_ok=True)
    image_path = os.path.join("uploads", image.filename)
    image.save(image_path)

    try:
        result = ocr_reader.read_image(image_path)
        return render_template("ocr_result.html", result=result)

    except Exception as e:
        flash(str(e), "error")
        return redirect(url_for("admin"))

    finally:
        if os.path.exists(image_path):
            os.remove(image_path)


@app.route("/admin/ocr_process", methods=["POST"])
@admin_required
def ocr_process():
    """Full process: OCR → compare by CRN/URN → download styled Excel."""
    image = request.files.get("manual_sheet")
    if not image or image.filename == "":
        flash("Please select an image", "error")
        return redirect(url_for("admin"))

    os.makedirs("uploads",       exist_ok=True)
    os.makedirs("excel_reports", exist_ok=True)

    image_path  = os.path.join("uploads",       image.filename)
    report_path = os.path.join("excel_reports", "ocr_final_report.xlsx")
    image.save(image_path)

    try:
        result          = ocr_reader.read_image(image_path)
        comparison_data = result["comparison_data"]

        if not comparison_data:
            flash("No matching records found. Check master Excel and image quality.", "error")
            return redirect(url_for("admin"))

        df            = pd.DataFrame(comparison_data)
        priority_cols = [c for c in ("CRN", "URN", "Status") if c in df.columns]
        other_cols    = [c for c in df.columns if c not in priority_cols]
        df            = df[priority_cols + other_cols]

        with pd.ExcelWriter(report_path, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Attendance")
            ws = writer.sheets["Attendance"]

            from openpyxl.styles import PatternFill, Font, Alignment

            green  = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            red    = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            header = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            bold_w = Font(bold=True, color="FFFFFF")

            for cell in ws[1]:
                cell.fill      = header
                cell.font      = bold_w
                cell.alignment = Alignment(horizontal="center")

            status_col_idx = next(
                (i for i, c in enumerate(ws[1], 1) if c.value == "Status"), None
            )

            if status_col_idx:
                for row in ws.iter_rows(min_row=2):
                    fill = green if row[status_col_idx - 1].value == "Present" else red
                    for cell in row:
                        cell.fill = fill

            for col_cells in ws.columns:
                length = max(len(str(c.value or "")) for c in col_cells)
                ws.column_dimensions[col_cells[0].column_letter].width = min(length + 4, 40)

        present_count = sum(1 for r in comparison_data if r.get("Status") == "Present")
        absent_count  = len(comparison_data) - present_count
        flash(f"Report ready — Present: {present_count} | Absent: {absent_count}", "success")

        return send_file(report_path, as_attachment=True, download_name="Final_Attendance_Status.xlsx")

    except Exception as e:
        import traceback
        traceback.print_exc()
        flash(f"Error processing image: {str(e)}", "error")
        return redirect(url_for("admin"))

    finally:
        if os.path.exists(image_path):
            os.remove(image_path)


# ── Analytics ─────────────────────────────────

@app.route("/analytics")
@admin_required
def analytics():
    conn = get_db_connection()

    branch_rows    = conn.execute("SELECT branch, COUNT(*) as count FROM attendance GROUP BY branch ORDER BY count DESC").fetchall()
    volunteer_rows = conn.execute("SELECT is_nss_volunteer, COUNT(*) as count FROM attendance GROUP BY is_nss_volunteer ORDER BY count DESC").fetchall()
    mode_rows      = conn.execute("SELECT attendance_mode, COUNT(*) as count FROM attendance GROUP BY attendance_mode ORDER BY count DESC").fetchall()
    event_rows     = conn.execute("SELECT event, COUNT(*) as count FROM attendance GROUP BY event ORDER BY count DESC").fetchall()

    conn.close()

    return render_template(
        "analytics.html",
        branch_data={
            "labels": [r["branch"] for r in branch_rows if r["branch"]],
            "counts": [r["count"]  for r in branch_rows if r["branch"]],
        },
        volunteer_data={
            "labels": [r["is_nss_volunteer"] for r in volunteer_rows if r["is_nss_volunteer"]],
            "counts": [r["count"]            for r in volunteer_rows if r["is_nss_volunteer"]],
        },
        mode_data={
            "labels": [r["attendance_mode"] for r in mode_rows if r["attendance_mode"]],
            "counts": [r["count"]           for r in mode_rows if r["attendance_mode"]],
        },
        event_data={
            "labels": [r["event"] for r in event_rows if r["event"]],
            "counts": [r["count"] for r in event_rows if r["event"]],
        },
    )


# ── Results / Download / API ──────────────────

@app.route("/results")
@admin_required
def results():
    conn    = get_db_connection()
    records = conn.execute("SELECT * FROM attendance ORDER BY id DESC").fetchall()
    conn.close()
    return render_template("results.html", records=records)


@app.route("/download")
@admin_required
def download():
    conn      = get_db_connection()
    file_path = export_attendance_to_excel(conn, "attendance_report.xlsx")
    conn.close()
    return send_file(file_path, as_attachment=True)


@app.route("/api/attendance")
@admin_required
def api_attendance():
    conn = get_db_connection()
    rows = conn.execute("SELECT * FROM attendance").fetchall()
    conn.close()
    return jsonify([dict(row) for row in rows])


@app.route("/api/locations")
@admin_required
def api_locations():
    """JSON feed of all attendance records that have GPS coordinates."""
    conn = get_db_connection()
    rows = conn.execute("""
        SELECT id, student_name, crn, urn, branch, section,
               event, attendance_mode, latitude, longitude, location_status
        FROM attendance
        WHERE latitude IS NOT NULL AND latitude != ''
    """).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


# ─────────────────────────────────────────────
#  Entry point
# ─────────────────────────────────────────────

if __name__ == "__main__":
    os.makedirs("static/qr_codes", exist_ok=True)
    os.makedirs("excel_reports",   exist_ok=True)
    os.makedirs("uploads",         exist_ok=True)
    migrate_db()   # adds latitude/longitude/location_status columns safely
    app.run(debug=True)
